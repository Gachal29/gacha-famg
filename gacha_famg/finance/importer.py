import datetime
from openpyxl import load_workbook

from db.models import Shop, Category, PurchaseRecordMonth, Receipt, PurchaseItem


class FinanceExcelImporter:
    home = None
    year_month = ""
    date = None
    filepath = None

    purchase_record_month_created = False
    shop_created = 0
    category_created = 0

    def __init__(self, home, year_month, date, filepath):
        self.home = home
        self.filepath = filepath
        self.year_month = year_month
        self.date = date

    def import_from_file(self):
        self.create_purchase_record_month()
        shop_list, categories, receipts = self.read_file()
        shop_objects = self.get_shop_objects(shop_list)
        category_objects = self.get_category_objects(categories)
        self.create_receipts(receipts, shop_objects, category_objects)
        return self.purchase_record_month_created, self.shop_created, self.category_created

    def create_purchase_record_month(self):
        obj, created = PurchaseRecordMonth.objects.update_or_create(
            home=self.home,
            year_month=self.year_month,
            defaults={
                "date": self.date
            },
            create_defaults={
                "home": self.home,
                "year_month": self.year_month,
                "date": self.date,
            }
        )
        self.purchase_record_month_created = created

    def read_file(self):
        workbook = load_workbook(self.filepath)
        # ショップリストの取得
        SHOP_NAME_COLUMN = "A"
        shop_list_ws = workbook["ショップリスト"]
        shop_num = shop_list_ws.max_row
        shop_list = []
        for row in range(1, shop_num+1):
            shop_list.append(shop_list_ws[f"{SHOP_NAME_COLUMN}{row}"].value)
        
        # 購入記録の取得
        DATE_COLUMN = "A"
        CATEGORY_COLUMN = "B"
        PRICE_COLUMN = "C"
        DETAIL_COLUMN = "D"
        SHOP_COLUMN = "E"

        purchase_records_ws = workbook["購入記録"]
        purchase_records_max_row = purchase_records_ws.max_row + 1
        purchase_records_start_row = 1
        for row  in range(1, purchase_records_max_row):
            if purchase_records_ws[f"{DATE_COLUMN}{row}"].value == "日付":
                purchase_records_start_row = row + 1
                break
        
        receipts = {}
        all_categories = []
        for row in range(purchase_records_start_row, purchase_records_max_row):
            date: datetime.datetime = purchase_records_ws[f"{DATE_COLUMN}{row}"].value
            if not isinstance(date, datetime.datetime):
                break

            date = date.date()
            key_date = str(date)
            try:
                receipts[key_date]
            except:
                receipts[key_date] = {}

            shop = purchase_records_ws[f"{SHOP_COLUMN}{row}"].value
            try:
                receipts[key_date][shop]
            except:
                receipts[key_date][shop] = []

            category = purchase_records_ws[f"{CATEGORY_COLUMN}{row}"].value
            price = purchase_records_ws[f"{PRICE_COLUMN}{row}"].value
            detail = purchase_records_ws[f"{DETAIL_COLUMN}{row}"].value

            if isinstance(price, str):
                price_formula = price[1:].split("*")
                price = int(int(price_formula[0]) * float(price_formula[1]))
            else:
                price = int(price)

            receipts[key_date][shop].append(dict(
                category=category,
                price=price,
                detail=detail,
            ))
            all_categories.append(category)
        categories = list(set(all_categories))
        return shop_list, categories, receipts
    
    def get_shop_objects(self, shop_list):
        shop_objects = {}
        for shop_name in shop_list:
            shop, created = Shop.objects.get_or_create(
                home=self.home,
                name=shop_name,
                defaults={
                    "home": self.home,
                    "name": shop_name,
                },
            )
            if created:
                self.shop_created += 1
            shop_objects[shop_name] = shop
        return shop_objects
    
    def get_category_objects(self, categories):
        category_objects = {}
        for category_name in categories:
            category, created = Category.objects.get_or_create(
                home=self.home,
                name=category_name,
                defaults={
                    "home": self.home,
                    "name": category_name,
                },
            )
            if created:
                self.category_created += 1
            category_objects[category_name] = category
        return category_objects
    
    def create_receipts(self, receipts_data, shop_objects, category_objects):
        receipt_objects = []
        purchase_item_objects = []
        receipt_days = receipts_data.keys()
        for date in receipt_days:
            receipt_shops = receipts_data[date].keys()
            for shop_name in receipt_shops:
                receipt_objects.append(Receipt(home=self.home, date=date, shop=shop_objects.get(shop_name)))
        
        receipts = Receipt.objects.bulk_create(receipt_objects)

        for receipt in receipts:
            for item in receipts_data[str(receipt.date)][receipt.shop.name]:
                purchase_item_objects.append(
                    PurchaseItem(
                        receipt=receipt,
                        category=category_objects.get(item.get("category")),
                        detail=item.get("detail"),
                        price=item.get("price"),
                    )
                )
        
        PurchaseItem.objects.bulk_create(purchase_item_objects)
